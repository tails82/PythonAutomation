__author__ = 'Tails'

from openpyxl import Workbook
from openpyxl import load_workbook
from testCases.TestCaseConfig import TestCaseConfig
import junit_xml
import common.nunit_xml
import sys
import traceback
import time
import os
import logging
import Logger

class Runner:

    def __init__(self):
        self.lstRunResult = []

    def run(self):
        self.currDateTime = time.strftime("%Y%m%d%H%M%S",time.localtime(time.time()))
        dicConfig = self.getConfigDic('config/config.xlsx')
        lstAllTestCaseConfig = self.getAllTestCases(dicConfig, 'config/Automation_Test_Suite.xlsx')
        Logger.logger = self.initailLogger(dicConfig)
        for testCaseConfig in lstAllTestCaseConfig:
            exec 'from testCases.%s.%s import %s' % (testCaseConfig.moduleName, testCaseConfig.testCaseID, testCaseConfig.testCaseID)
            testCaseInstance = eval(testCaseConfig.testCaseID)(dicConfig, testCaseConfig)
            try:
                testCaseInstance.setUp()
                testCaseInstance.run()
            except Exception as ex:
                #save screen shot when failed
                testCaseInstance.saveScreenShot()
                testCaseInstance.allPassed = False
                testCaseConfig.testCaseRunResult = "Failed"
                traceback.print_exc(file=sys.stdout)
                testCaseConfig.failureStackTrace = traceback.format_exc()
                testCaseConfig.failureMessage = ex.message
            finally:
                testCaseInstance.tearDown()
                self.saveTestResult(testCaseConfig)
        self.outputResult(testCaseInstance)

    def getConfigDic(self, fileName):
         wb = load_workbook(filename=fileName, read_only=True)
         #get which env is selected
         envWs = wb.get_sheet_by_name('Env')
         dicEnvConfig = {}
         rowId = 0
         for row in envWs.rows:
             #skip the first row, it's title
             if rowId == 0:
                 rowId += 1
                 continue
             lstCurrCell = []
             for cell in row:
                 lstCurrCell.append(cell.value)
             #insert the config item to the dictionary
             item = lstCurrCell[0]
             value = lstCurrCell[1]
             dicEnvConfig[item] = value
             rowId += 1
         env = dicEnvConfig['Environment']
         dicConfig = {}
         #read all configs in the [env] sheet
         configWs = wb.get_sheet_by_name(env)
         rowId = 0
         for row in configWs.rows:
             #skip the first row, it's title
             if rowId == 0:
                 rowId += 1
                 continue
             lstCurrCell = []
             for cell in row:
                 lstCurrCell.append(cell.value)
             #insert the config item to the dictionary
             item = lstCurrCell[0]
             value = str(lstCurrCell[1])
             dicConfig[item] = value
             rowId += 1
         return dicConfig

    def getAllTestCases(self, dicConfig, fileName):
        """
        :param fileName:
        :return: A list contains all testcase (testCaseConfig) to be run
        """
        lstAllTestCases = []
        lstAllTestCaseConfig = []
        wb = load_workbook(filename=fileName, read_only=True)
        sheetName = dicConfig['Test Suite']
        lstAllTestCases = []
        ws = wb[sheetName]
        currTestCase = []
        for row in ws.rows:
            currTestCase = []
            currTestCase.append(sheetName)
            for cell in row:
                currTestCase.append(cell.value)
            if currTestCase[3] == 'Run':
                lstAllTestCases.append(currTestCase)
        for testCase in lstAllTestCases:
            currTestCaseConfig = TestCaseConfig(testCase[0], testCase[1], testCase[2])  #sheetName, moduleName, testCaseID
            lstAllTestCaseConfig.append(currTestCaseConfig)
        return lstAllTestCaseConfig

    def saveTestResult(self, testCaseConfig):
        tupleRunResult = (testCaseConfig.sheetName, testCaseConfig.moduleName, testCaseConfig.testCaseID, testCaseConfig.testCaseRunResult,
                          testCaseConfig.timeElapsedSec, testCaseConfig.failureMessage, testCaseConfig.failureStackTrace)
        self.lstRunResult.append(tupleRunResult)

    def outputResult(self, testCaseInstance):
        """
        Write test case run result(pass, failed) to excel file
        :param lstTestResult:
        :return:
        """
        #create a folder with timestamp to save current test run result
        currDir = os.getcwd()
        resultDir = currDir + os.path.sep + testCaseInstance.dicConfig['Test Result Folder']
        runResultDir = resultDir + os.path.sep +  self.currDateTime
        os.mkdir(runResultDir)
        self.generateExcelReport(self.lstRunResult, runResultDir)
        self.generateJUnitReport(self.lstRunResult, resultDir)
        self.generateNunitReport(self.lstRunResult, resultDir)

    def generateExcelReport(self, lstRunResult, runResultDir):
         #create excel file and write test result in
        resultFileName = runResultDir + os.path.sep + 'RunResult.xlsx'
        wb = Workbook()
        previousCaseSheetName = ''
        rowIndex = 0
        for runResult in lstRunResult:
            #runResult (sheetName, moduleName, testCaseID, runResult, timeElapsedSec, failureMessage, failureStackTrace)
            currCaseModuleName = runResult[0]
            #first sheet
            if previousCaseSheetName == "":
                ws = wb.active
                ws.title = currCaseModuleName
            #start another sheet
            elif not previousCaseSheetName == currCaseModuleName:
                ws = wb.create_sheet(title=currCaseModuleName)
                #reset row index in a new sheet
                rowIndex = 0
            rowIndex += 1
            ws['A' + str(rowIndex)] = runResult[1]
            ws['B' + str(rowIndex)] = runResult[2]
            ws['C' + str(rowIndex)] = runResult[3]
            previousCaseSheetName = currCaseModuleName
        wb.save(filename=resultFileName)

    def generateJUnitReport(self, lstRunResult, runResultDir):
        #create junit xml report file use junit-xml 1.4   pip install junit-xml
        resultFileName = runResultDir + os.path.sep + 'RunResult.xml'
        previousCaseModuleName = ''
        rowIndex = 0
        lstTestSuites = []
        testSuite = []
        for runResult in lstRunResult:
            #runResult (sheetName, moduleName, testCaseID, runResult, timeElapsedSec, failureMessage, failureStackTrace)
            #test
            testCaseName = runResult[2]
            className = runResult[1] + '.' + runResult[2]
            timeElapsedSec = runResult[4]
            failureMessage = runResult[5]
            failureStackTrace = runResult[6]
            testCase = junit_xml.TestCase(testCaseName, className, timeElapsedSec)
            testCase.add_failure_info(None, failureMessage)
            currTestCaseModuleName = runResult[1]
            if not currTestCaseModuleName == previousCaseModuleName:
                testSuite = junit_xml.TestSuite(currTestCaseModuleName)
                lstTestSuites.append(testSuite)
            testSuite.test_cases.append(testCase)
        #print TestSuite.to_xml_string(lstTestSuites)
        #Write the xml content to result file
        with open(runResultDir + os.path.sep + 'JUnitResult.xml', 'w') as f:
            junit_xml.TestSuite.to_file(f, lstTestSuites)

    def generateNunitReport(self, lstRunResult, runResultDir):
        resultFileName = runResultDir + os.path.sep + 'RunResult.xml'
        previousCaseModuleName = ''
        rowIndex = 0
        lstTestSuites = []
        testSuite = []
        for runResult in lstRunResult:
            #runResult (sheetName, moduleName, testCaseID, runResult, timeElapsedSec, failureMessage, failureStackTrace)
            #test
            testCaseName = runResult[2]
            time = runResult[4]
            failureMessage = runResult[5]
            failureStackTrace = runResult[6]
            testCase = common.nunit_xml.TestCase(testCaseName, time)
            testCase.add_failure_info(message=failureMessage, stackTrace=failureStackTrace)
            currTestCaseModuleName = runResult[1]
            if not currTestCaseModuleName == previousCaseModuleName:
                testSuite = common.nunit_xml.TestSuite(currTestCaseModuleName)
                lstTestSuites.append(testSuite)
            testSuite.test_cases.append(testCase)
        #print common.nunit_xml.TestSuite.to_xml_string(lstTestSuites)
        #Write the xml content to result file
        with open(runResultDir + os.path.sep + 'NUnitResult.xml', 'w') as f:
            common.nunit_xml.TestSuite.to_file(f, lstTestSuites)

    def initailLogger(self, dicConfig):
        currDateTime = time.strftime("%Y%m%d%H%M%S",time.localtime(time.time()))
        currDir = os.getcwd()
        resultDir = currDir + os.path.sep + dicConfig['Log Folder']
        logFileName = resultDir + os.path.sep + "Log" + currDateTime + ".log"
        logging.basicConfig(level=logging.INFO,
                            format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                            datefmt='%a, %d %b %Y %H:%M:%S',
                            filename=logFileName,
                            filemode='w')
        logger = logging.getLogger()
        return logging.getLogger()